from openpyxl import load_workbook
import time

def xl_input():
    print("excel 데이터 가져오는중...")
    file_name = "address.xlsx"
    wb = load_workbook(file_name)
    ws = wb["1. 총괄표(현행)"]
    address = []
    for row in range(4, 3795):
        if ws.cell(row=row, column=7).value != None:
            list = [ws.cell(row=row, column=3).value, ws.cell(row=row, column=5).value, ws.cell(row=row, column=7).value]
            address.append(list)
    return address